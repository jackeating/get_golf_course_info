from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
from datetime import datetime

def get_url_list(driver):
    url_list = []
    for page_num in range(1, 14):
        driver.get("https://golf-medley.com/practice_facilities/prefectures/kanagawa?page=" + str(page_num))
        elems = driver.find_elements_by_xpath("//*[@id='__next']/*/*/*/*/*/*/a")
        for elem in elems:
            href_data = elem.get_attribute("href")
            if "reviews" in href_data:
                continue
            if "prefectures" in href_data:
                continue
            url_list.append(href_data)
    return url_list

def main():
    result_list = []
    dame_list = []
    error_dict = {
        "ryoukin": "料金テーブル習得不可: ",
        "name": "施設名称の取得不可: ",
        "time": "営業時間等の取得不可: ",
    }
    driver = webdriver.Chrome(executable_path="./chromedriver")
    try:

        url_list = get_url_list(driver)
        num = 0
        for url in url_list:
            num += 1
            driver.get(url)
            try:
                name = driver.find_element_by_xpath("//*[@id='__next']/*/*/*/span[4]").text
            except:
                print(error_dict["name"] + url)
                continue
            if "閉鎖" in name:
                continue
            result_dict = {
                "施設名称": name
            }
            
            # 営業時間、アクセス、支払い方法の取得
            try:
                for x_path in (
                    "//*[@id='__next']/div[2]/div[4]/div[1]/div[1]/div/div/table",
                    "//*[@id='__next']/div[2]/div[4]/div[1]/div[2]/div/div[1]/table",
                    "//*[@id='__next']/div[2]/div[4]/div[2]/div/div/div[2]/table"
                    ):
                    tableElem = driver.find_element_by_xpath(x_path)
                    trs = tableElem.find_elements(By.TAG_NAME, "tr")
                    for i in range(0,len(trs)):
                        tds = trs[i].find_elements(By.TAG_NAME, "td")
                        result_dict.update({
                            tds[0].text: tds[1].text
                        })
            except:
                print(error_dict["time"] + url)
                dame_list.append(url)
                continue

            try:
                pay = driver.find_element_by_xpath("//*[@id='__next']/div[2]/div[4]/div[5]/div[1]/div/div/table/tbody/tr[8]/td[2]").text
            except:
                pay = ""
            result_dict.update({
                "決済方法": pay
            })

            # 料金の取得
            try:

                priceTable = driver.find_element_by_xpath("//*[@id='__next']/div[2]/div[4]/div[2]/div/div/div[1]/table")
                trs = priceTable.find_elements(By.TAG_NAME, "tr")
                for i in range(1,len(trs)):
                    th = trs[i].find_elements(By.TAG_NAME, "th")
                    tds = trs[i].find_elements(By.TAG_NAME, "td")
                    column_name = th[0].text
                    # 平日の処理
                    result_dict.update({
                        column_name + "(平日)": tds[0].text
                    })
                    # 土日祝の処理
                    result_dict.update({
                        column_name + "(土日祝)": tds[1].text
                    })
            except:
                print(error_dict["ryoukin"] + url)
            
            try:
                rental_club = driver.find_element_by_xpath("//*[@id='__next']/div[2]/div[4]/div[5]/div[1]/div/div/table/tbody/tr[9]/td[2]").text
            except:
                rental_club = ""
            
            try:
                home_page = driver.find_element_by_xpath("//*[@id='__next']/div[2]/div[4]/div[5]/div[2]/div/div[1]/table/tbody/tr[2]/td[2]/a").text
            except:
                home_page = ""

            result_dict.update({
                "レンタルクラブ": driver.find_element_by_xpath("//*[@id='__next']/div[2]/div[4]/div[3]/div/div/div[12]/div/p[2]").text,
                "レンタルクラブ(レッスン情報)": rental_club,
                "ホームページ": home_page,
            })
            result_list.append(result_dict)

    except:
        driver.quit()
    if len(result_list) == 0:
        return

    column_list = result_list[0].keys()
    wb=openpyxl.Workbook()
    ws=wb.active
    # シート名の設定
    ws.title="ゴルフ打ちっぱなし"
    # シートの読み込み
    sheet = wb['ゴルフ打ちっぱなし']
    # 値の代入
    for number, column_data in enumerate(column_list, start=1):
        sheet.cell(column=number,row=1).value = column_data

    row_num = 2
    for row_data in result_list:
        for column_number, column_data in enumerate(column_list, start=1):
            sheet.cell(column=column_number,row=row_num).value = row_data.get(column_data, "")
        row_num += 1
    today = datetime.today()
    wb.save('./{}年{}月_data.xlsx'.format(str(today.year), str(today.month)))
if __name__ == '__main__':
    main()